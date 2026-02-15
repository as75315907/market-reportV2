import os
import json
import math
import time
import re
import smtplib
from datetime import datetime
from email.message import EmailMessage
import subprocess
import certifi

# 可選：使用作業系統信任庫（對某些公司網路/防毒 HTTPS 攔截更穩）
try:
    import truststore  # pip install truststore
    truststore.inject_into_ssl()
except Exception:
    pass
import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook
from dotenv import load_dotenv
from tenacity import retry, stop_after_attempt, wait_exponential

# ========= 基本設定 =========
TEMPLATE_XLSX = "每日股價行情表.xlsx"   # 你的「想要的版面」模板
OUTPUT_PREFIX = "Daily_Market_Report"

SHEET_NAME = "IR_updated (PC HOME)"     # 你截圖那個 sheet
CACHE_FILE = "tw_suffix_cache.json"     # 台股 .TW / .TWO 快取（自動判斷用）

# 指數（Yahoo symbols）
TICKER_TWII = "^TWII"
TICKER_HSI  = "^HSI"

# 台股成交額（官方）
TWSE_MI_INDEX = "https://www.twse.com.tw/exchangeReport/MI_INDEX"

# ========= SSL / requests Session（最重要） =========
CA_BUNDLE = certifi.where()
os.environ.setdefault("SSL_CERT_FILE", CA_BUNDLE)
os.environ.setdefault("REQUESTS_CA_BUNDLE", CA_BUNDLE)

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
})
SESSION.verify = CA_BUNDLE
SESSION.trust_env = False  # 不讀 HTTPS_PROXY / 代理設定，避免走到被MITM的proxy


# ========= 小工具 =========
def _isnan(x) -> bool:
    try:
        return x is None or (isinstance(x, float) and (math.isnan(x) or math.isinf(x)))
    except Exception:
        return True

def _round2(x):
    if _isnan(x):
        return None
    try:
        return round(float(x), 2)
    except Exception:
        return None

def load_cache() -> dict:
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_cache(cache: dict):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ========= Yahoo 價格 =========
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
def hist_one(ticker: str) -> pd.DataFrame:
    # 用 1mo 讓「最近兩個交易日」更穩
    return yf.Ticker(ticker).history(period="1mo", interval="1d", auto_adjust=False)

def has_enough_prices(hist: pd.DataFrame) -> bool:
    if hist is None or hist.empty:
        return False
    if "Close" not in hist.columns:
        return False
    return hist["Close"].dropna().shape[0] >= 2

def resolve_tw_ticker(code: str, cache: dict) -> str:
    """
    台股 ticker 常見兩種：XXXX.TW（上市）/ XXXX.TWO（上櫃）
    這裡自動測試、成功就寫入 cache。
    """
    code = str(code).strip()
    if code in cache:
        return f"{code}.{cache[code]}"

    for suf in ["TWO", "TW"]:
        t = f"{code}.{suf}"
        try:
            h = hist_one(t)
            if has_enough_prices(h):
                cache[code] = suf
                return t
        except Exception:
            continue

    cache[code] = "TW"
    return f"{code}.TW"

def hk_ticker(code: str) -> str:
    # Yahoo 港股 ticker：四碼 + .HK，例如 03368 -> 3368.HK；00825 -> 0825.HK
    return f"{int(str(code)):04d}.HK"

def last_two(series: pd.Series):
    s = series.dropna()
    if len(s) < 2:
        return (pd.NaT, math.nan, pd.NaT, math.nan)
    return s.index[-1], float(s.iloc[-1]), s.index[-2], float(s.iloc[-2])

def read_hk_lot_map() -> dict:
    """
    讀 hk_lot.csv（可選）
    格式：code,lot_size
    """
    mp = {}
    fn = "hk_lot.csv"
    if not os.path.exists(fn):
        return mp
    try:
        df = pd.read_csv(fn, dtype=str).fillna("")
        for _, r in df.iterrows():
            code = str(r.get("code", "")).strip()
            lot  = str(r.get("lot_size", "")).strip().replace(",", "")
            if code and lot.isdigit():
                mp[code.replace(".HK", "").zfill(4)] = int(lot)
    except Exception:
        return mp
    return mp

def ps_get_text(url: str) -> str:
    """
    用 PowerShell Invoke-WebRequest 取得網頁文字（走 Windows Schannel，通常可避開 OpenSSL 的 SKI 問題）
    """
    cmd = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy", "Bypass",
        "-Command",
        f"(Invoke-WebRequest -UseBasicParsing -Uri '{url}').Content"
    ]
    p = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if p.returncode != 0:
        # 有些環境會把錯誤寫在 stderr
        raise RuntimeError(f"PowerShell Invoke-WebRequest 失敗: {p.stderr.strip()}")
    return p.stdout

# ========= Email =========


def curl_get_text(url: str) -> str:
    """用系統 curl 下載（Windows 通常內建 curl.exe），可避開 Python/OpenSSL 驗證差異。"""
    # -L 跟隨 redirect；--compressed 接受壓縮；--ssl-no-revoke 避免部分 Windows 憑證撤銷檢查問題
    cmd = ["curl", "-L", "--compressed", "--ssl-no-revoke", "-A", "Mozilla/5.0", url]
    try:
        return subprocess.check_output(cmd, text=True, encoding="utf-8", errors="ignore")
    except FileNotFoundError:
        # 沒有 curl 的環境（較少見）
        raise
    except subprocess.CalledProcessError as e:
        raise RuntimeError(e.stderr or str(e))

def fetch_text_fallback(url: str) -> str:
    """依序嘗試 requests -> curl -> powershell。"""
    try:
        r = SESSION.get(url, timeout=30)
        r.raise_for_status()
        return r.text
    except Exception:
        pass

    # 2) curl
    try:
        return curl_get_text(url)
    except Exception:
        pass

    # 3) PowerShell（最後手段；仍走系統憑證）
    return ps_get_text(url)


def send_email_with_attachment(subject, body, attachment_path):
    load_dotenv()
    host = os.getenv("SMTP_HOST")
    port = int(os.getenv("SMTP_PORT", "465"))
    user = os.getenv("SMTP_USER")
    app_pw = os.getenv("SMTP_APP_PASSWORD")
    mail_to = os.getenv("MAIL_TO")

    if not all([host, port, user, app_pw, mail_to]):
        raise RuntimeError("請先在 .env 設定 SMTP_HOST/SMTP_PORT/SMTP_USER/SMTP_APP_PASSWORD/MAIL_TO")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = user
    msg["To"] = mail_to
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        data = f.read()

    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(attachment_path),
    )

    with smtplib.SMTP_SSL(host, port) as smtp:
        smtp.login(user, app_pw)
        smtp.send_message(msg)


# ========= 讀模板，找「台股/港股」列 =========
def find_stock_rows_from_template(ws):
    """
    依模板版面自動找：
    - 台股區：從「台股（台幣）」那行下面開始，直到 column A 空白為止
    - 港股區：從「港股（港幣）」那行下面開始，直到 column A 空白為止
    回傳：
    tw_rows = [(row, code_str), ...]
    hk_rows = [(row, code_str), ...]
    """
    tw_start = None
    hk_start = None

    for r in range(1, 200):
        v = ws[f"B{r}"].value
        if isinstance(v, str) and "台股" in v and "台幣" in v:
            tw_start = r + 1
        if isinstance(v, str) and "港股" in v and "港幣" in v:
            hk_start = r + 1

    if tw_start is None or hk_start is None:
        raise RuntimeError("找不到模板中的『台股（台幣）/港股（港幣）』標題列，請確認模板 sheet 是否正確。")

    tw_rows = []
    r = tw_start
    while r < hk_start:
        code = ws[f"A{r}"].value
        if code is None or str(code).strip() == "":
            break
        code_s = str(code).strip()
        if code_s.isdigit() and len(code_s) == 4:
            tw_rows.append((r, code_s))
        r += 1

    hk_rows = []
    r = hk_start
    for _ in range(0, 50):
        code = ws[f"A{r}"].value
        if code is None or str(code).strip() == "":
            break
        code_s = str(code).strip()
        code_s = code_s.replace(".HK", "").replace("HK", "").strip()
        hk_rows.append((r, code_s))
        r += 1

    return tw_rows, hk_rows


# ========= 抓價：把需要的欄位整理成 dict =========
def build_ohlcv_map(ticker_list: list[str]) -> dict:
    """
    回傳 dict:
    {
      "2926.TWO": {
        "t_date":..., "p_date":...,
        "close":..., "prev_close":..., "open":..., "high":..., "low":..., "volume":...
      },
      ...
    }
    """
    out = {}
    for t in ticker_list:
        try:
            h = hist_one(t)
        except Exception:
            h = pd.DataFrame()

        if h is None or h.empty or "Close" not in h.columns or h["Close"].dropna().shape[0] < 2:
            out[t] = {}
            continue

        t_date, t_close, p_date, p_close = last_two(h["Close"])

        def _last(col):
            if col not in h.columns:
                return math.nan
            s = h[col].dropna()
            return float(s.iloc[-1]) if len(s) else math.nan

        out[t] = {
            "t_date": t_date,
            "p_date": p_date,
            "close": t_close,
            "prev_close": p_close,
            "open": _last("Open"),
            "high": _last("High"),
            "low":  _last("Low"),
            "volume": _last("Volume"),
        }
        time.sleep(0.2)
    return out


# ========= TWSE 成交額（最穩、最準） =========
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
def fetch_twse_total_turnover_value(trade_date: datetime) -> int:
    """
    從 TWSE 官方 MI_INDEX 抓「總計」成交金額（新台幣元）。
    成功率最高的做法是：
      1) requests (verify=certifi)；
      2) 若 SSL 在你環境仍失敗，改用系統 curl / PowerShell 下載同一頁面再解析。
    回傳 int，例如 699292298413
    """
    yyyymmdd = trade_date.strftime("%Y%m%d")
    url = f"{TWSE_MI_INDEX}?response=html&date={yyyymmdd}&type=ALL"

    html = fetch_text_fallback(url)

    # 先用 pandas.read_html 解析表格（比 regex 穩）
    try:
        tables = pd.read_html(html)
        for df in tables:
            if df is None or df.empty or df.shape[1] < 2:
                continue

            # 找「總計」那列
            col0 = df.iloc[:, 0].astype(str)
            hit = col0.str.contains("總計") | col0.str.contains("Total", case=False, na=False)
            if not bool(hit.any()):
                continue

            row = df[hit].iloc[0]

            # 優先找欄名含「成交金額/成交值/Turnover」
            col_names = [str(c) for c in df.columns]
            target_idx = None
            for j, cn in enumerate(col_names):
                if any(k in cn for k in ["成交金額", "成交值", "Turnover", "Trade Value"]):
                    target_idx = j
                    break

            def _to_int(x):
                if x is None:
                    return None
                s = str(x)
                s = re.sub(r"[^0-9]", "", s)
                return int(s) if s else None

            if target_idx is not None:
                v = _to_int(row.iloc[target_idx])
                if v is not None:
                    return v

            # 沒找到欄名就用「最大數字」當成交金額（通常最大的是成交金額）
            nums = []
            for j in range(1, df.shape[1]):
                v = _to_int(row.iloc[j])
                if v is not None:
                    nums.append(v)
            if nums:
                return max(nums)
    except Exception:
        pass

    # 最後才用 regex（以免表格結構變動）
    m = re.search(r"總計[^0-9]*([0-9,]{8,})", html)
    if not m:
        raise RuntimeError("MI_INDEX 找不到『總計』成交金額（可能頁面結構更新/被擋）")
    return int(m.group(1).replace(",", ""))

def to_yi(x: int) -> float:
    # 轉「億」單位（1 億 = 100,000,000）
    return round(x / 100_000_000, 2)


# ========= 寫入模板固定格（不動格式/公式） =========
def write_cell(ws, addr: str, value):
    ws[addr].value = value


def main():
    if not os.path.exists(TEMPLATE_XLSX):
        raise FileNotFoundError(f"找不到模板檔：{TEMPLATE_XLSX}（請把它放在同一資料夾）")

    wb = load_workbook(TEMPLATE_XLSX)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"模板中找不到工作表：{SHEET_NAME}")

    ws = wb[SHEET_NAME]

    # 依模板找出台股/港股行
    tw_rows, hk_rows = find_stock_rows_from_template(ws)

    cache = load_cache()

    tw_codes = [code for _, code in tw_rows]
    hk_codes = []
    for _, code in hk_rows:
        c = str(code).strip().replace(".HK", "")
        # 港股代號通常 4-5 碼，yahoo 要 4 碼（0825.HK / 3368.HK）
        hk_codes.append(c.zfill(4))

    tw_tickers = [resolve_tw_ticker(c, cache) for c in tw_codes]
    hk_tickers = [hk_ticker(c) for c in hk_codes]
    save_cache(cache)

    # 抓價（個股 + 指數）
    stock_map = build_ohlcv_map(tw_tickers + hk_tickers)
    idx_map   = build_ohlcv_map([TICKER_TWII, TICKER_HSI])

    # 寫日期（你模板 L3）
    now = datetime.now()
    write_cell(ws, "L3", now)

    # ===== 指數：填指數 & 成交額 =====
    twii = idx_map.get(TICKER_TWII, {})
    hsi  = idx_map.get(TICKER_HSI, {})

    # 指數（依你原本位置 D6/E6，D8/E8）
    write_cell(ws, "D6", _round2(twii.get("close")))
    write_cell(ws, "E6", _round2(twii.get("prev_close")))
    write_cell(ws, "D8", _round2(hsi.get("close")))
    write_cell(ws, "E8", _round2(hsi.get("prev_close")))

    # ===== TWII 成交額（H6=今日, I6=前一日；單位：十億台幣）=====
    tw_today_dt = twii.get("t_date")
    tw_prev_dt  = twii.get("p_date")

    if isinstance(tw_today_dt, pd.Timestamp):
        tw_today_dt = tw_today_dt.to_pydatetime()
    if isinstance(tw_prev_dt, pd.Timestamp):
        tw_prev_dt = tw_prev_dt.to_pydatetime()

    tw_today_bil = None
    tw_prev_bil = None
    if tw_today_dt and tw_prev_dt:
        try:
            tw_today_val = fetch_twse_total_turnover_value(tw_today_dt)
            tw_prev_val  = fetch_twse_total_turnover_value(tw_prev_dt)
            tw_today_bil = to_yi(tw_today_val)
            tw_prev_bil  = to_yi(tw_prev_val)
        except Exception as e:
            print("WARN: 台股成交額抓取失敗，將留白。原因：", e)


    write_cell(ws, "H6", tw_today_bil)
    write_cell(ws, "I6", tw_prev_bil)

    # ===== HSI 成交額 =====
    # 你先前已能顯示 HSI 成交額，這份先「不覆蓋」模板 H8/I8
    # 如果你希望我也把 HSI 成交額改成同樣「官方來源 + 兩交易日」邏輯，我再幫你接 HKEX 或穩定替代源。

    # ===== 港股手數：讀 hk_lot.csv，沒有就預設 1000 股/手 =====
    hk_lot = read_hk_lot_map()

    # ===== 台股區（只寫：D/E/H/I/J/K；不動公式欄 F/G/L）=====
    for (r, code), ticker in zip(tw_rows, tw_tickers):
        d = stock_map.get(ticker, {})
        close = _round2(d.get("close"))
        prev  = _round2(d.get("prev_close"))
        opn   = _round2(d.get("open"))
        low   = _round2(d.get("low"))
        high  = _round2(d.get("high"))
        vol   = d.get("volume")

        lots = None
        if not _isnan(vol):
            lots = int(round(float(vol) / 1000))

        write_cell(ws, f"D{r}", close)
        write_cell(ws, f"E{r}", prev)
        write_cell(ws, f"H{r}", opn)
        write_cell(ws, f"I{r}", low)
        write_cell(ws, f"J{r}", high)
        write_cell(ws, f"K{r}", lots)

    # ===== 港股區（成交張數用「手數」）=====
    for (r, raw_code), code, ticker in zip(hk_rows, hk_codes, hk_tickers):
        d = stock_map.get(ticker, {})
        close = _round2(d.get("close"))
        prev  = _round2(d.get("prev_close"))
        opn   = _round2(d.get("open"))
        low   = _round2(d.get("low"))
        high  = _round2(d.get("high"))
        vol   = d.get("volume")

        lot_size = hk_lot.get(code.zfill(4), 1000)
        hands = None
        if not _isnan(vol) and lot_size and lot_size > 0:
            hands = int(round(float(vol) / lot_size))

        write_cell(ws, f"D{r}", close)
        write_cell(ws, f"E{r}", prev)
        write_cell(ws, f"H{r}", opn)
        write_cell(ws, f"I{r}", low)
        write_cell(ws, f"J{r}", high)
        write_cell(ws, f"K{r}", hands)

    # 另存新檔（避免你開著 Excel 造成 PermissionError）
    today = now.strftime("%Y-%m-%d")
    stamp = now.strftime("%H%M%S")
    out_file = f"{OUTPUT_PREFIX}_{today}_{stamp}.xlsx"
    wb.save(out_file)

    # 寄信
    subject = f"每日股市整理_{today}_17-30"
    body = (
        "附上今日股市整理（版型=每日股價行情表.xlsx 模板）。\n\n"
        "本版更新：\n"
        "- ✅ 台股 TWII 成交額：採 TWSE 官方 MI_INDEX（總計成交金額），自動抓最近兩交易日並換算成『十億台幣』寫入 H6/I6。\n"
        "- 台股成交張數 = Yahoo Volume / 1000。\n"
        "- 港股成交張數 = 手數（Volume / lot_size），lot_size 讀 hk_lot.csv，未提供則預設 1000。\n"
        "- HSI 成交額目前先不覆蓋模板既有值（你原先已可顯示）。\n"
    )
    send_email_with_attachment(subject, body, out_file)

    print("DONE:", out_file)
    print(f"TW rows: {len(tw_rows)} | HK rows: {len(hk_rows)}")
    print(f"TWII turnover (today/prev, 十億台幣): {tw_today_bil} / {tw_prev_bil}")


if __name__ == "__main__":
    main()