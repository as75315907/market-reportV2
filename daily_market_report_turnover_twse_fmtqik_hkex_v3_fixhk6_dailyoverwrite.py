# -*- coding: utf-8 -*-
"""
Daily Market Report (Template-based)
- Prices: yfinance (stocks + indices)
- TWII turnover value: TWSE FMTQIK JSON (成交金額(元)) -> 億元
- Robust TLS on Windows: try truststore (system cert store) first
- Keeps your Excel template layout untouched; only writes values to fixed cells.
"""

import os
import json
import math
import time
import smtplib
import re
import subprocess
from pathlib import Path
from datetime import datetime, timedelta
from email.message import EmailMessage
from io import StringIO

import pandas as pd
import yfinance as yf
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv
from tenacity import retry, stop_after_attempt, wait_exponential

# ========= Path Base =========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# ====== (Optional) Use Windows system cert store for better success on TWSE ======
# pip install truststore
try:
    import truststore  # type: ignore
    truststore.inject_into_ssl()
except Exception:
    pass

# ========= 基本設定 =========
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

TEMPLATE_XLSX = os.path.join(BASE_DIR, "每日股價行情表.xlsx")   # 你的「想要的版面」模板
OUTPUT_PREFIX = "Daily_Market_Report"
SHEET_NAME = "IR_updated (PC HOME)"     # 你截圖那個 sheet
CACHE_FILE = os.path.join(BASE_DIR, "tw_suffix_cache.json")     # 台股 .TW / .TWO 快取（自動判斷用）
DEBUG_HKEX = os.getenv("DEBUG_HKEX", "") == "1"
DEBUG_DIR = os.path.join(BASE_DIR, "debug")
def _debug_save(name: str, text: str):
    if not DEBUG_HKEX:
        return
    try:
        os.makedirs(DEBUG_DIR, exist_ok=True)
        Path(os.path.join(DEBUG_DIR, name)).write_text(text or "", encoding="utf-8", errors="ignore")
    except Exception:
        pass
     # 台股 .TW / .TWO 快取（自動判斷用）

# 指數（Yahoo symbols）
TICKER_TWII = "^TWII"
TICKER_HSI  = "^HSI"

# TWSE 月資料：成交金額(元) 來源（JSON）
TWSE_FMTQIK = "https://www.twse.com.tw/exchangeReport/FMTQIK"

# ========= 小工具 =========
def _isnan(x) -> bool:
    try:
        return x is None or (isinstance(x, float) and (math.isnan(x) or math.isinf(x)))
    except Exception:
        return True

def _round2(x):
    if _isnan(x):
        return None
    try:
        return round(float(x), 2)
    except Exception:
        return None

def load_cache() -> dict:
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_cache(cache: dict):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# yfinance occasionally throttles; keep conservative retries
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
def hist_one(ticker: str) -> pd.DataFrame:
    return yf.Ticker(ticker).history(period="1mo", interval="1d", auto_adjust=False)

def has_enough_prices(hist: pd.DataFrame) -> bool:
    if hist is None or hist.empty:
        return False
    if "Close" not in hist.columns:
        return False
    return hist["Close"].dropna().shape[0] >= 2

def resolve_tw_ticker(code: str, cache: dict) -> str:
    """
    台股 ticker 常見：XXXX.TW（上市）/ XXXX.TWO（上櫃）
    自動測試可用者並寫入 cache。
    """
    code = str(code).strip()
    if code in cache:
        return f"{code}.{cache[code]}"

    for suf in ["TW", "TWO"]:
        t = f"{code}.{suf}"
        try:
            h = hist_one(t)
            if has_enough_prices(h):
                cache[code] = suf
                return t
        except Exception:
            continue

    cache[code] = "TW"
    return f"{code}.TW"

def hk_ticker(code: str) -> str:
    return f"{int(str(code)):04d}.HK"

def last_two(series: pd.Series):
    s = series.dropna()
    if len(s) < 2:
        return (pd.NaT, math.nan, pd.NaT, math.nan)
    return s.index[-1], float(s.iloc[-1]), s.index[-2], float(s.iloc[-2])

def read_hk_lot_map() -> dict:
    """
    讀 hk_lot.csv（可選）
    格式：code,lot_size
    """
    mp = {}
    fn = os.path.join(BASE_DIR, "hk_lot.csv")
    if not os.path.exists(fn):
        return mp
    try:
        df = pd.read_csv(fn, dtype=str).fillna("")
        for _, r in df.iterrows():
            code = str(r.get("code", "")).strip().replace(".HK", "").strip()
            lot  = str(r.get("lot_size", "")).strip().replace(",", "")
            if code and lot.isdigit():
                mp[code.zfill(4)] = int(lot)
    except Exception:
        return mp
    return mp

# ========= 寄信 =========
def send_email_with_attachment(subject, body, attachment_path):
    """
    本機可用 .env；GitHub Actions 用環境變數/Secrets。
    若 SMTP 相關環境變數不齊，就直接跳過寄信（不讓整個流程失敗）。
    """
    # 本機若有 .env 才載入；Actions 沒 .env 也不會出錯
    try:
        load_dotenv()
    except Exception:
        pass

    host = os.getenv("SMTP_HOST", "").strip()
    port_s = os.getenv("SMTP_PORT", "465").strip()
    user = os.getenv("SMTP_USER", "").strip()
    app_pw = os.getenv("SMTP_APP_PASSWORD", "").strip()
    mail_to = os.getenv("MAIL_TO", "").strip()

    # Actions 沒設定就跳過（避免 exit code 1）
    if not (host and port_s and user and app_pw and mail_to):
        print("[INFO] SMTP env not set. Skip sending email.")
        return

    try:
        port = int(port_s)
    except Exception:
        port = 465

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = user
    msg["To"] = mail_to
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        data = f.read()

    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(attachment_path),
    )

    with smtplib.SMTP_SSL(host, port) as smtp:
        smtp.login(user, app_pw)
        smtp.send_message(msg)

# ========= 讀模板，找「台股/港股」列 =========
def find_stock_rows_from_template(ws):
    """
    依模板版面自動找：
    - 台股區：從「台股（台幣）」那行下面開始，直到 column A 空白為止
    - 港股區：從「港股（港幣）」那行下面開始，直到 column A 空白為止
    """
    tw_start = None
    hk_start = None

    for r in range(1, 300):
        v = ws[f"B{r}"].value
        if isinstance(v, str) and ("台股" in v and "台幣" in v):
            tw_start = r + 1
        if isinstance(v, str) and ("港股" in v and "港幣" in v):
            hk_start = r + 1

    if tw_start is None or hk_start is None:
        raise RuntimeError("找不到模板中的『台股（台幣）/港股（港幣）』標題列，請確認模板 sheet 是否正確。")

    tw_rows = []
    r = tw_start
    while r < hk_start:
        code = ws[f"A{r}"].value
        if code is None or str(code).strip() == "":
            break
        code_s = str(code).strip()
        if code_s.isdigit():
            tw_rows.append((r, code_s))
        r += 1

    hk_rows = []
    r = hk_start
    for _ in range(0, 60):
        code = ws[f"A{r}"].value
        if code is None or str(code).strip() == "":
            break
        code_s = str(code).strip()
        code_s = code_s.replace(".HK", "").replace("HK", "").strip()
        hk_rows.append((r, code_s))
        r += 1

    return tw_rows, hk_rows

# ========= 抓價：把需要的欄位整理成 dict =========
def build_ohlcv_map(ticker_list: list[str]) -> dict:
    """
    回傳 dict:
    { ticker: {t_date, p_date, close, prev_close, open, high, low, volume} }
    """
    out = {}
    for t in ticker_list:
        try:
            h = hist_one(t)
        except Exception:
            h = pd.DataFrame()

        if h is None or h.empty or "Close" not in h.columns or h["Close"].dropna().shape[0] < 2:
            out[t] = {}
            continue

        t_date, t_close, p_date, p_close = last_two(h["Close"])

        def _last(col):
            if col not in h.columns:
                return math.nan
            s = h[col].dropna()
            return float(s.iloc[-1]) if len(s) else math.nan

        out[t] = {
            "t_date": t_date,
            "p_date": p_date,
            "close": float(t_close),
            "prev_close": float(p_close),
            "open": _last("Open"),
            "high": _last("High"),
            "low":  _last("Low"),
            "volume": _last("Volume"),
        }
        time.sleep(0.2)
    return out

def write_cell(ws, addr: str, value):
    ws[addr].value = value

# ========= TWSE 成交金額（元） -> 億元 =========
_SESSION = requests.Session()
_SESSION.headers.update({
    "User-Agent": UA,
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
    "Referer": "https://www.twse.com.tw/zh/trading/historical/fmtqik.html",
})

def _safe_get(url: str, *, timeout: int = 30, headers: dict | None = None, verify: bool | None = None):
    """優先用 verify=True；若遇到公司網路 SSL 攔截導致驗證失敗，會自動降級 verify=False。
    這是為了讓抓取不中斷，同時在 DEBUG_HKEX=1 時把錯誤寫入 debug/。
    """
    hdrs = {}
    if headers:
        hdrs.update(headers)
    # explicit verify if caller set it
    if verify is not None:
        return _SESSION.get(url, timeout=timeout, headers=hdrs, verify=verify)

    try:
        return _SESSION.get(url, timeout=timeout, headers=hdrs, verify=True)
    except requests.exceptions.SSLError as e:
        _debug_save("ssl_error.log", f"{url}\n{repr(e)}\n")
        return _SESSION.get(url, timeout=timeout, headers=hdrs, verify=False)

# ========= 港股成交額（HK 市場成交額，寫入 HSI 那一列）來源：HKEX Daily Quotations =========
HKEX_DAYQUOT = "https://www.hkex.com.hk/eng/stat/smstat/dayquot/d{yymmdd}e.htm"
HKEX_DAYQUOT_REFERER = "https://www.hkex.com.hk/eng/stat/smstat/dayquot/qtn.asp"

# AASTOCKS fallback (顯示「成交額◎ X億」)
AASTOCKS_HSI_URL = "https://www.aastocks.com/tc/stocks/market/index/hk-index-con.aspx?index=HSI&o=0&p=&s=8&t=6"

def _curl_get_text(url: str, timeout: int = 30, insecure: bool = False, http1: bool = False, extra_headers: list[str] | None = None) -> str:
    """curl 後備方案（避免某些 Windows/SSL 環境問題）。
    insecure=True 會加 -k（略過憑證驗證），只在 verify 失敗時最後手段使用。
    http1=True 會加 --http1.1（HKEX 某些情況下 HTTP/2 會回空頁/導覽頁）。
    extra_headers: e.g. ["Referer: ...", "Accept-Language: ..."]
    """
    cmd = ["curl", "-L", "-s", "--max-time", str(timeout), "-A", UA]
    if http1:
        cmd.append("--http1.1")
    if extra_headers:
        for h in extra_headers:
            cmd += ["-H", h]
    if insecure:
        cmd.insert(1, "-k")
    cmd.append(url)
    return subprocess.check_output(cmd, text=True, encoding="utf-8", errors="ignore")


def _hkex_yymmdd(dt: datetime | None) -> str:
    if dt is None or pd.isna(dt):
        return ""
    if isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime()
    if isinstance(dt, datetime):
        return dt.strftime("%y%m%d")
    return ""

def fetch_hkex_dayquot_html(trade_dt: datetime) -> str:
    yymmdd = _hkex_yymmdd(trade_dt)
    if not yymmdd:
        raise RuntimeError("HKEX 日期解析失敗")
    url = HKEX_DAYQUOT.format(yymmdd=yymmdd)

    # 先用 requests（配合 truststore，通常在公司網路最穩）
    try:
        r = _safe_get(
            url,
            timeout=30,
            verify=True,
            headers={
                "User-Agent": UA,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9,zh-TW;q=0.8,zh;q=0.7",
                "Referer": HKEX_DAYQUOT_REFERER,
            },
        )
        r.raise_for_status()
        # 讓 requests 用偵測到的編碼解碼（HKEX 有時不是 utf-8）
        try:
            r.encoding = r.apparent_encoding or r.encoding
        except Exception:
            pass
        html = r.text or ""
        _debug_save(f"hkex_dayquot_{yymmdd}_requests.html", html)

        # 簡單 sanity check：避免抓到導覽/空頁
        low = html.lower()
        if ("turnover" not in low) and ("成交" not in html) and ("daily quotations" not in low):
            raise RuntimeError("HKEX 回傳內容疑似非日報頁（缺少 turnover 關鍵字）")
        return html
    except Exception as e_req:
        # 再用 curl（有的環境 requests 會被 SSL 攔）
        curl_headers = [
            f"Referer: {HKEX_DAYQUOT_REFERER}",
            "Accept-Language: en-US,en;q=0.9,zh-TW;q=0.8,zh;q=0.7",
        ]
        try:
            html = _curl_get_text(url, timeout=30, insecure=False, http1=True, extra_headers=curl_headers)
            _debug_save(f"hkex_dayquot_{yymmdd}_curl.html", html)
            return html
        except Exception:
            html = _curl_get_text(url, timeout=30, insecure=True, http1=True, extra_headers=curl_headers)
            _debug_save(f"hkex_dayquot_{yymmdd}_curl_insecure.html", html)
            return html


def parse_hkex_turnover_hkd(html: str) -> float:
    """
    從 HKEX Day Quotations HTML 抓出『成交額』，回傳 **港幣金額（HKD）**。

    典型表格會有欄位名稱：
    - "Turnover (HK$ Million)" 或 "Turnover (HK$)"（也可能是中文：成交額/成交金額）
    我們會優先從欄位名稱判斷單位，再抓「Total/總計」或最後一列的數值。
    """
    if not html:
        raise RuntimeError("HKEX HTML 空白")

    # ---- Fast path: regex directly on raw HTML (more robust than read_html when HKEX changes table layout) ----
    # Typical labels:
    #   "Total Market Turnover (HK$ Million)"  => value is in HK$ million
    #   Sometimes Chinese label may appear as well.
    rx_candidates = [
        r"Total\s+Market\s+Turnover\s*\(\s*HK\$\s*Million\s*\)[^0-9]{0,300}([0-9][0-9,]*\.?[0-9]*)",
        r"Total\s+Market\s+Turnover[^0-9]{0,300}([0-9][0-9,]*\.?[0-9]*)",
        r"市場\s*成交額[^0-9]{0,300}([0-9][0-9,]*\.?[0-9]*)",
        r"成交\s*(?:額|金額)[^0-9]{0,300}([0-9][0-9,]*\.?[0-9]*)",
    ]
    for rx in rx_candidates:
        m = re.search(rx, html, flags=re.I)
        if m:
            v = float(m.group(1).replace(",", ""))
            # HKEX dayquot is almost always HK$ million; treat as million when value looks like 10,000+.
            if v >= 10_000:
                return v * 1_000_000.0
            # Otherwise keep as HKD (best-effort)
            return v

    try:
        tables = pd.read_html(StringIO(html))
    except Exception as e:
        raise RuntimeError(f"HKEX read_html 失敗：{e}")

    def _to_float(x):
        s = str(x).strip().replace(",", "").replace("\u00a0", " ")
        m = re.search(r"([0-9][0-9]*\.?[0-9]*)", s)
        if not m:
            return None
        try:
            return float(m.group(1))
        except Exception:
            return None

    def _unit_multiplier(col_text: str, sample_val: float | None):
        t = (col_text or "").lower()
        if "million" in t or "mn" in t:
            return 1_000_000.0
        if "billion" in t or "bn" in t:
            return 1_000_000_000.0
        if "百萬" in col_text or "百万" in col_text:
            return 1_000_000.0
        if "十億" in col_text or "十亿" in col_text:
            return 1_000_000_000.0
        # 未標示：HKEX 常見數值為 HK$ Mn
        if sample_val is not None and 10_000 <= sample_val <= 1_000_000:
            return 1_000_000.0
        return 1.0

    for df in tables:
        try:
            df = df.fillna("")
        except Exception:
            pass

        cols = [str(c).strip() for c in getattr(df, "columns", [])]
        if not cols:
            continue

        turnover_cols = []
        for i, c in enumerate(cols):
            cl = c.lower()
            if ("turnover" in cl) or ("成交額" in c) or ("成交金額" in c) or ("成交金额" in c):
                turnover_cols.append(i)

        if not turnover_cols:
            # fallback：逐列找包含 turnover/成交額 的列
            for _, row in df.iterrows():
                cells = [str(x).strip() for x in row.tolist()]
                joined = " | ".join(cells)
                low = joined.lower()
                if ("turnover" in low) or ("成交額" in joined) or ("成交金額" in joined) or ("成交金额" in joined):
                    for c in cells:
                        v = _to_float(c)
                        if v is not None:
                            mult = _unit_multiplier(joined, v)
                            return v * mult
            continue

        idx = turnover_cols[0]
        col_name = cols[idx]
        series = df.iloc[:, idx]

        pick_val = None
        if df.shape[1] >= 2:
            first_col = df.iloc[:, 0].astype(str).str.strip()
            mask = first_col.str.contains(r"^(total|總計|合計)$", case=False, regex=True)
            if mask.any():
                v = _to_float(series[mask].iloc[-1])
                if v is not None:
                    pick_val = v

        if pick_val is None:
            vals = []
            for x in series.tolist():
                v = _to_float(x)
                if v is not None:
                    vals.append(v)
            if vals:
                pick_val = vals[-1]

        if pick_val is None:
            continue

        mult = _unit_multiplier(col_name, pick_val)
        return pick_val * mult

    m = re.search(r"turnover[^0-9]{0,40}([0-9][0-9,]*\.?[0-9]*)", html, flags=re.I)
    if m:
        v = float(m.group(1).replace(",", ""))
        mult = 1_000_000.0 if 10_000 <= v <= 1_000_000 else 1.0
        return v * mult

    raise RuntimeError("HKEX 找不到 Turnover 欄位")



# ========= AASTOCKS fallback =========

def fetch_aastocks_hsi_html() -> str:
    r = _safe_get(
        AASTOCKS_HSI_URL,
        timeout=30,
        headers={
            "User-Agent": UA,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
            "Referer": "https://www.aastocks.com/",
        },
    )
    r.raise_for_status()
    try:
        r.encoding = r.apparent_encoding or r.encoding
    except Exception:
        pass
    html = r.text or ""
    _debug_save("aastocks_hsi.html", html)
    return html

def parse_aastocks_turnover_yi(html: str) -> float:
    """
    回傳成交額（**億港幣**，即 HKD 1e8）。
    AASTOCKS 可能出現的格式：
      - 中文頁：成交額◎ 2,575.78億
      - 英文頁：Turnover◎ 257.58B  (B=HKD billions)
      - 也可能在 <span class="turnover">...</span> 內
    """
    if not html:
        raise RuntimeError("AASTOCKS HTML 空白")

    # 1) Most robust: the dedicated turnover span (works even when label is split across tags)
    m = re.search(r'class=["\']turnover["\'][^>]*>\s*([0-9][0-9,]*\.?[0-9]*)\s*([BbMm億])', html)
    if m:
        num = float(m.group(1).replace(",", ""))
        unit = m.group(2)
        if unit in ("億",):
            return round(num, 2)
        if unit in ("B", "b"):
            # HKD billions -> /1e8 = *10
            return round(num * 10.0, 2)
        if unit in ("M", "m"):
            # HKD millions -> /1e8 = /100
            return round(num / 100.0, 2)

    # 2) Fallback: label-based
    m = re.search(r'(?:成交額|Turnover)[^0-9]{0,50}([0-9][0-9,]*\.?[0-9]*)\s*([BbMm億])', html, flags=re.I)
    if m:
        num = float(m.group(1).replace(",", ""))
        unit = m.group(2)
        if unit in ("億",):
            return round(num, 2)
        if unit in ("B", "b"):
            return round(num * 10.0, 2)
        if unit in ("M", "m"):
            return round(num / 100.0, 2)

    # 3) Last resort: find a "turnover ... 億" anywhere (some pages show plain text)
    m = re.search(r"turnover[^0-9]{0,80}([0-9][0-9,]*\.?[0-9]*)\s*([BbMm億])", html, flags=re.I)
    if m:
        num = float(m.group(1).replace(",", ""))
        unit = m.group(2)
        if unit == "億":
            return round(num, 2)
        if unit in ("B", "b"):
            return round(num * 10.0, 2)
        if unit in ("M", "m"):
            return round(num / 100.0, 2)

    raise RuntimeError("AASTOCKS 找不到成交額/Turnover")


def fetch_hkex_turnover_yi(date_dt: datetime) -> float:
    """成交額（億港幣）。"""
    html = fetch_hkex_dayquot_html(date_dt)
    hkd = parse_hkex_turnover_hkd(html)
    return round(hkd / 1e8, 2)

def get_two_hkex_turnover_by_hsi_dates(hsi_today_dt, hsi_prev_dt) -> tuple[float | None, float | None]:
    """用 yfinance 的 ^HSI 交易日對齊 HKEX Turnover。HKEX 失敗時，今日值會改用 AASTOCKS 作備援。"""
    out_today = None
    out_prev = None

    # ---- today ----
    try:
        if hsi_today_dt is not None and not pd.isna(hsi_today_dt):
            d = hsi_today_dt.to_pydatetime() if isinstance(hsi_today_dt, pd.Timestamp) else hsi_today_dt
            out_today = fetch_hkex_turnover_yi(d)
    except Exception:
        out_today = None

    if out_today is None:
        # AASTOCKS 只能保證「最新交易日」成交額；用來避免 H8 全空
        try:
            html = fetch_aastocks_hsi_html()
            out_today = round(parse_aastocks_turnover_yi(html), 2)
        except Exception:
            out_today = None

    # ---- prev ----
    try:
        if hsi_prev_dt is not None and not pd.isna(hsi_prev_dt):
            d = hsi_prev_dt.to_pydatetime() if isinstance(hsi_prev_dt, pd.Timestamp) else hsi_prev_dt
            out_prev = fetch_hkex_turnover_yi(d)
    except Exception:
        out_prev = None

    return out_today, out_prev


def _ad_to_twse_date_str(dt: datetime) -> str:
    # TWSE FMTQIK data row uses ROC year like "115/02/11"
    roc_y = dt.year - 1911
    return f"{roc_y:03d}/{dt.month:02d}/{dt.day:02d}"

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=8))
def fetch_fmtqik_month_json(dt: datetime) -> dict:
    # FMTQIK uses date=YYYYMM01 (month key)
    month_key = dt.strftime("%Y%m") + "01"
    params = {"response": "json", "date": month_key}
    r = _SESSION.get(TWSE_FMTQIK, params=params, timeout=30)
    r.raise_for_status()
    obj = r.json()
    stat = str(obj.get("stat", ""))
    if "沒有" in stat or "No data" in stat:
        raise RuntimeError(f"FMTQIK no data for {month_key}")
    return obj

def extract_turnover_from_fmtqik(obj: dict, dt: datetime) -> int | None:
    """
    Return 成交金額(元) as int, or None.
    We match by ROC date string in first column.
    """
    target = _ad_to_twse_date_str(dt)
    data = obj.get("data", [])
    if not isinstance(data, list):
        return None

    # Find column index for 成交金額 (header may vary slightly)
    fields = obj.get("fields", [])
    col_idx = None
    if isinstance(fields, list):
        for i, f in enumerate(fields):
            fs = str(f)
            if "成交金額" in fs:
                col_idx = i
                break

    for row in data:
        if not isinstance(row, list) or not row:
            continue
        if str(row[0]).strip() != target:
            continue

        # if header matched
        if col_idx is not None and col_idx < len(row):
            s = str(row[col_idx]).replace(",", "").strip()
            if s.isdigit():
                return int(s)

        # fallback: last numeric-like field in row is usually amount
        for cell in reversed(row):
            s = str(cell).replace(",", "").strip()
            if s.isdigit():
                return int(s)
    return None

def twse_turnover_yi(dt: datetime) -> float | None:
    """
    Return turnover in 億元 (TWD 1e8).
    """
    try:
        obj = fetch_fmtqik_month_json(dt)
        val = extract_turnover_from_fmtqik(obj, dt)
        if val is None:
            return None
        return round(val / 1e8, 2)
    except Exception:
        return None


def normalize_hk_turnover_to_yi(val):
    """
    把港股成交額數值正規化為「億港幣」。

    因資料來源可能回傳：
    - 已是「億港幣」（例如 2575.78）
    - HK$ million（百萬港幣），若誤乘/誤解析會變成 2,575,775,925.24 這種超大值
    - 其他倍數誤差（例如多乘 1,000 或 1,000,000）

    這個函式用「合理區間」做兜底縮放：
    - 港股全市場單日成交額通常落在 50 ~ 50,000 億港幣（極端日可能更高，但很少超過此量級）
    """
    if val is None:
        return None
    try:
        x = float(val)
    except Exception:
        return None

    x = abs(x) if x < 0 else x

    # 已在合理範圍
    if 0 <= x <= 50000:
        return x

    # 太小：可能是以「萬億/兆」或其他單位表示，這邊不做放大，避免誤判
    if x < 0.01:
        return x

    # 太大：常見是多了 1,000 或 1,000,000
    # 先嘗試除以 1,000,000（最常見：把『億』又乘回『百萬』）
    for _ in range(6):
        if x <= 50000:
            break
        if x >= 1e7:
            x = x / 1e6
        else:
            x = x / 1e3
    return x

def main():
    if not os.path.exists(TEMPLATE_XLSX):
        raise FileNotFoundError(f"找不到模板檔：{TEMPLATE_XLSX}（請把它放在同一資料夾）")

    wb = load_workbook(TEMPLATE_XLSX)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"模板中找不到工作表：{SHEET_NAME}")
    ws = wb[SHEET_NAME]

    tw_rows, hk_rows = find_stock_rows_from_template(ws)

    cache = load_cache()
    tw_codes = [code for _, code in tw_rows]

    hk_codes = []
    for _, code in hk_rows:
        c = str(code).strip().replace(".HK", "")
        # keep as 4-digit for hk_ticker
        c = c.zfill(4)
        hk_codes.append(c)

    tw_tickers = [resolve_tw_ticker(c, cache) for c in tw_codes]
    hk_tickers = [hk_ticker(c) for c in hk_codes]
    save_cache(cache)

    stock_map = build_ohlcv_map(tw_tickers + hk_tickers)
    idx_map   = build_ohlcv_map([TICKER_TWII, TICKER_HSI])

    now = datetime.now()
    write_cell(ws, "L3", now)

    # ===== Indices =====
    twii = idx_map.get(TICKER_TWII, {})
    hsi  = idx_map.get(TICKER_HSI, {})

    write_cell(ws, "D6", _round2(twii.get("close")))
    write_cell(ws, "E6", _round2(twii.get("prev_close")))

    # TWII turnover: use TWSE FMTQIK by the same trade dates as Yahoo (best consistency)
    tw_t_date = twii.get("t_date")
    tw_p_date = None
    # Derive prev date from yfinance history again (safer than reading from map)
    try:
        h = hist_one(TICKER_TWII)
        t_date, _, p_date, _ = last_two(h["Close"])
        tw_t_date, tw_p_date = t_date, p_date
    except Exception:
        pass

    tw_today_yi = None
    tw_prev_yi  = None
    if isinstance(tw_t_date, pd.Timestamp):
        tw_today_yi = twse_turnover_yi(tw_t_date.to_pydatetime())
    if isinstance(tw_p_date, pd.Timestamp):
        tw_prev_yi = twse_turnover_yi(tw_p_date.to_pydatetime())

    # H6/I6 顯示億元
    write_cell(ws, "H6", tw_today_yi)
    write_cell(ws, "I6", tw_prev_yi)

    # HSI index close/prev (成交額你目前能顯示就維持原邏輯；此版本不改 H8/I8)
    write_cell(ws, "D8", _round2(hsi.get("close")))
    write_cell(ws, "E8", _round2(hsi.get("prev_close")))
    # 港股市場成交額（億港元）→ 寫入 H8/I8
    hk_today_yi, hk_prev_yi = get_two_hkex_turnover_by_hsi_dates(hsi.get("t_date"), hsi.get("p_date"))
    hk_today_yi = normalize_hk_turnover_to_yi(hk_today_yi)
    hk_prev_yi  = normalize_hk_turnover_to_yi(hk_prev_yi)
    write_cell(ws, "H8", hk_today_yi)
    write_cell(ws, "I8", hk_prev_yi)

    # ===== Stocks =====
    hk_lot = read_hk_lot_map()

    # 台股區（不動公式欄）
    for (r, code), ticker in zip(tw_rows, tw_tickers):
        d = stock_map.get(ticker, {})
        close = _round2(d.get("close"))
        prev  = _round2(d.get("prev_close"))
        opn   = _round2(d.get("open"))
        low   = _round2(d.get("low"))
        high  = _round2(d.get("high"))
        vol   = d.get("volume")

        lots = None
        if not _isnan(vol):
            lots = int(round(float(vol) / 1000))

        write_cell(ws, f"D{r}", close)
        write_cell(ws, f"E{r}", prev)
        write_cell(ws, f"H{r}", opn)
        write_cell(ws, f"I{r}", low)
        write_cell(ws, f"J{r}", high)
        write_cell(ws, f"K{r}", lots)

    # 港股區（成交張數=手數）
    for (r, raw_code), code, ticker in zip(hk_rows, hk_codes, hk_tickers):
        d = stock_map.get(ticker, {})
        close = _round2(d.get("close"))
        prev  = _round2(d.get("prev_close"))
        opn   = _round2(d.get("open"))
        low   = _round2(d.get("low"))
        high  = _round2(d.get("high"))
        vol   = d.get("volume")

        lot_size = hk_lot.get(code.zfill(4), 1000)
        hands = None
        if not _isnan(vol) and lot_size and lot_size > 0:
            hands = int(round(float(vol) / lot_size))

        write_cell(ws, f"D{r}", close)
        write_cell(ws, f"E{r}", prev)
        write_cell(ws, f"H{r}", opn)
        write_cell(ws, f"I{r}", low)
        write_cell(ws, f"J{r}", high)
        write_cell(ws, f"K{r}", hands)

    today = now.strftime("%Y-%m-%d")
    # ===== Output =====
    # 同日覆蓋：檔名固定只含日期；預設輸出到 reports/（可用環境變數 OUTPUT_DIR 改路徑）
    output_dir = os.path.join(BASE_DIR, os.getenv("OUTPUT_DIR", "reports"))
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"{OUTPUT_PREFIX}_{today}.xlsx")
    wb.save(out_file)

    subject = f"每日股市整理_{today}_17-30"
    body = (
        "附上今日股市整理（模板版型）。\n\n"
        "更新：\n"
        "- TWII 成交額：改用 TWSE FMTQIK(JSON) 的『成交金額(元)』，轉為『億元』寫入 H6/I6。\n"
        "- 若遇到公司網路/防毒 SSL 攔截，建議安裝 truststore：py -m pip install truststore\n"
    )
    send_email_with_attachment(subject, body, out_file)

    print("DONE:", out_file)
    print(f"TW rows: {len(tw_rows)} | HK rows: {len(hk_rows)}")
    print(f"TWII turnover (today/prev, 億元): {tw_today_yi} / {tw_prev_yi}")
    print(f"HK turnover (today/prev, 億港幣): {hk_today_yi} / {hk_prev_yi}")

if __name__ == "__main__":
    main()
